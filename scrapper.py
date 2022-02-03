#!/usr/bin/env python
# encoding: utf-8

'''FOURTH COMMIT: Just change the filepath (excel file) of the filename to the path in your system and Run this script.
    I'm building a shape for the tables.

    I'm doing it with specific values but later it will be for any output value.

'''

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
from openpyxl.styles import Font, NamedStyle

from openpyxl.utils.cell import coordinate_from_string, get_column_letter
from openpyxl.utils.cell import get_column_letter
import os

# ========================================================================================================
# Exploring data
# ========================================================================================================

# We export the data to a pre config file later:

'''Just change the filepath of the filename and Run this script.'''

filename = r"D:\Arquivos HD\Projetos HD\SD Labs\transformer\SPSS_output_scrapper-master\WEIGHTED_customer_database.xlsx"

# Import your dataset, for example:
wb = openpyxl.load_workbook(filename)
# worksheet active:
ws = wb.active

# is a dictionary-like object of all the tables in a particular worksheet:
ws.tables

# tables = ws.tables: Returns a list of table name and their ranges.
ws.tables.items()


# ========================================================================================================
# Class to parser and handle spps outputs
# ========================================================================================================

class SpssOutputScrapper():

    def __init__(self):
        super().__init__()

    def crawl_spss_output(self):
        pass

    def get_table_name(self):
        name = ws.tables[self.table_name]
        return name

    def getTableColumn(self):
        '''
        Crawl the first rows of the table to identify the labels
        Stop to crawl when background is grey

        '''
        self.table = ws.tables

        table_range = ws.table.ref

    def getTableRow(self):
        '''
                if rowType == 'Total':
                # Crawl the first column of the
                elif rowType == 'Count'
                # Identify valid values in second column & identify
                elif rowType == 'Prev'

        # Identify count and add 1 row (Prev/Perc always comes after count)
        '''
        pass

    def identifyTableEnd(self):

        '''
        # Identify "total" in first column.
        # Or identify change i cell formattings (white vs grey background)
        '''
        pass

    def get_table_name(self, table_name):
        name = ws.tables[table_name]
        return name

    def find_table(self, table_name, tables):
        '''
            # Use: table_range = find_table(table_name, tables)

            Parameters
            ----------
            table_name: string
                Name of the Table.
            tables : string

            Returns
            -------

        '''

        for table in tables:
            if table.displayName == table_name:
                return table.ref

    def table_info(self):

        for table in ws.tables.values():
            print(table.headerRowCount)
            print(table.ref)
            print(len(table.tableColumns))
            print(table)

    def delete_table(self, table_name):
        del ws.tables[table_name]


# ========================================================================================================
# Class to Extract data of the unstructured output file.
# ========================================================================================================

class ExtractTables:

    def __init__(self):
        super().__init__()

    def extract_title(self):
        pass

    def extract_columns(self):
        pass

    def extract_rows(self):
        pass

    pass


# ========================================================================================================
# Class to Feed the Table structure for the extraction.
# ========================================================================================================

class Keywords():
    pass


# ========================================================================================================
# Exploring DataFrame
# ========================================================================================================

'''
excel = pd.read_excel(filename)

categories = [
    x for x in excel.columns if excel[x].dtype == 'object' #or x == 'inadimplente'
    ]

print(categories)

for v in categories:
    print('\n%15s:' % v, "%4d categorias" % len(excel[v].unique()))
    print(excel[v].unique(), '\n')

'''

# At final, the scripts bellow will give all the positions needed to structure and accesses of the Tables. v

# We can pass the values to a predefined config file later.

sheet = wb["Sheet1"]

predefined_table_name_list = ['$Var_set*agecat Crosstabulation',
                              '$Var_set Frequencies',
                              'Notes',
                              'Case Processing Summary',
                              'Gender * Age category Crosstabulation',
                              'Union member * Age category Crosstabulation',
                              'Retired * Age category Crosstabulation',
                              'Marital status * Age category Crosstabulation'
                              ]

extract_table_name_list = [
    'Gender * Age category Crosstabulation',
    'Union member * Age category Crosstabulation',
    'Retired * Age category Crosstabulation',
    'Marital status * Age category Crosstabulation'
]

categories_table_name_list = ['Gender',
                              'Total',
                              'Union member',
                              'Retired',
                              'Marital status']

sub_table_name_list = ['Male',
                       'Female',
                       'No',
                       'yes',
                       'Unmarried',
                       'Married',

                       ]

sub_categories_table_name_list = []

counts = ['Count']

counts_percentages = ['% within Union member',
                      '% within Age category',
                      '% within Retired',
                      '% within Marital status']

age_category = ['18-24',
                '25-34',
                '35-49',
                '50-64',
                '>65'
                ]

title_age_category = ['Age category']

print('predefined table name list \n')

for i in sheet['A1':'A161']:  # section for all table name available.
    for n in i:
        if n.value in predefined_table_name_list:
            print(n.coordinate, n.value)

print('\n')

print('extract table name list\n')
for i in sheet['A1':'A161']:  # section to extraction of table names for the new tables.
    for n in i:
        if n.value in extract_table_name_list:
            print(n.coordinate, n.value)

print('\n')

print('categories table name list\n')
for i in sheet['A1':'A161']:  # section to extraction of categories for the new tables.
    for n in i:
        if n.value in categories_table_name_list:
            print(n.coordinate, n.value)

print('\n')

print('sub table name list\n')
for i in sheet['B1':'B161']:  # section to extraction of table sub names for the new tables.
    for n in i:
        if n.value in sub_table_name_list:
            print(n.coordinate, n.value)

print('\n')

print('sub categories table name list\n')
for i in sheet['A1':'C161']:  # section to extraction of sub categories for the new tables.
    for n in i:
        if n.value in sub_categories_table_name_list:
            print(n.coordinate, n.value)

print('\n')

print('counts \n')
for i in sheet['C1':'C161']:  # section
    for n in i:
        if n.value in counts:
            print(n.coordinate, n.value)

print('\n')

print('counts percentages \n')
for i in sheet['C1':'C161']:  # section
    for n in i:
        if n.value in counts_percentages:
            print(n.coordinate, n.value)

print('\n')

print('age category \n')
for i in sheet['D1':'I161']:  # section
    for n in i:
        if n.value in age_category:
            print(n.coordinate, n.value)

print('\n')

print(' title age category \n')
for i in sheet['D1':'H161']:  # section
    for n in i:
        if n.value in title_age_category:
            print(n.coordinate, n.value)

# I Found the positions of all items of the tables
# now I will using this positions to construct the table and turn it accessible .

# ========================================================================================================
# Structuring all the tables.
# ========================================================================================================


# The number of tables to know how many tables we will create.
number_of_tables = len(extract_table_name_list)

sheet = wb["Sheet1"]
counter = 0  # count how many tables will be create.
coordinates = []
num_row = []
table_size = []

for i in sheet['A1':'A161']:  # read the row in sheet - The Table title appears in the A column
    for n in i:  # read the cell in sheet
        if n.value in extract_table_name_list:
            coor = n.coordinate
            # print('coordinate')
            # print(coor)
            coordinates.append(coor)

            row_number = n.row
            # print('row number')
            # print(row_number)
            # print(type(row_number))
            num_row.append(row_number)

            val = n.value
            # print('value')
            # print(val)

            # table_name_used.append(num)
            # access the line above the next extract_table_name_list
            n = 1  # 0 is the first title
            # first get the next title
            next_title = extract_table_name_list[n]
            # print(next_title)
            # print(type(next_title))
            n + 1  # rise one index number

            # print('index')
            # print(extract_table_name_list.index(next_title))

        break
    counter = + 1

    if counter == number_of_tables:
        print("all the tables were created")
        break

print('Title rows: {}'.format(num_row))
print('\n')
# print('Title coordinates: {}'.format(coordinates))

# End tables by blank space

# print('table per rows')
end_table = []
end_tables = []
spaces = []
i = 0
f = 1
for j in range(len(num_row)):
    end_table = num_row[f] - num_row[i]
    spaces.append(end_table)
    i + 1
    f + 1
# print(spaces) # lenght of table rows

# print('end table')


zip_object = zip(num_row, spaces)
for num_row, spaces in zip_object:
    end_tables.append(num_row + spaces)
end_tables = [x - 1 for x in end_tables]  # one line above the table title is blank

new_coord = []
print('using openpyxl.utils.cell.coordinate_from_string(')
for i in coordinates:
    a, b = coordinate_from_string(i)
    new_coord.append(b)

print('Title coordinates: {}'.format(new_coord))
print('this is the ends of each of the tables: ')
print(end_tables)

# Get the end table coordinate

print('End Tables coordinates')
end_tables_coordinate = [('A' + str(i)) for i in end_tables]
print(end_tables_coordinate)

# plot the tables by section:

workbook = openpyxl.Workbook()
worksheet = workbook.active
columns_to_copy = [('M' + str(i)) for i in end_tables]
positions = 0
coordinate_to_copy = []
positions2 = 0
letter = ''
positions3 = 0

# Create the coordinates for each section
# Use to copy the worksheets

for i in range(int(len(new_coord[positions2])), int(len(end_table[positions2]))):

    while positions3 < 20:  # Convert a column index into a column letter (3 -> ‘C’)
        att = get_column_letter(positions3) + str(i)
        coordinate_to_copy.append(att)
        positions3 + 1

    positions2 + 1
    if positions2 == number_of_tables:
        break




print('plot the tables by section:')
for i in sheet[coordinates[positions]:end_tables_coordinate[positions]]:

    # Try to write each table per column

    for n in i:

        worksheet["A1"] = sheet[["A1"]]

        if n.value in extract_table_name_list:
            print(n.coordinate, n.value)

    positions + 1
    if positions == number_of_tables:
        break

# here i will verify each row sections collected from A to I column

# define type of table
def table_type_1():
    pass


# ========================================================================================================
# Create a new structured Table with the data scrapped from the output
# ========================================================================================================

'''
# create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# populate some sample data
# I will improve this part to the pythonic method later
# just to show that the table will be accessible.
# we gonna to automate the inputs.

worksheet["A1"] = "Gender * Age category Crosstabulation"
worksheet["B1"] = " "
worksheet["A2"] = " "
worksheet["A3"] = " "
worksheet["B2"] = "Age Category"
worksheet["B3"] = " "
worksheet["C3"] = " "
worksheet["D3"] = "18-24"
worksheet["E3"] = "25-34"
worksheet["F3"] = "35-49"
worksheet["G3"] = "50-64"
worksheet["H3"] = ">65"
worksheet["I3"] = "Total"

worksheet["A4"] = "Gender"
worksheet["B4"] = "Male"
worksheet["B7"] = "Female"
worksheet["A10"] = "Total"

# define a table style
mediumStyle = TableStyleInfo(name='TableStyleMedium2',
                             showRowStripes=True)
# create a table
table = openpyxl.worksheet.table.Table(ref='A1:I10',
                                       displayName='Gender_*_Age_category_Crosstabulation',
                                       tableStyleInfo=mediumStyle)
# add the table to the worksheet
worksheet.add_table(table)

# save the workbook file
workbook.save('tables.xlsx')'''
